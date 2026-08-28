#!/usr/bin/env python3
"""Build a clean, ready-to-use AttackCoverage.xlsx for MITRE ATT&CK v19.2.

The pristine upstream v11 workbook (20220505/AttackCoverage.xlsx) is used as
the base so that every formula, cell style, differential style, conditional
formatting rule and table definition is inherited verbatim.  Only three kinds
of thing change:

  * the data rows (techniques / tactics / sources, regenerated from the v19.2
    STIX bundle by get_tt.py),
  * the sheet extents (697 techniques instead of 578, 872 tactic rows instead
    of 750, 98 data sources instead of 109),
  * the tactic list -- Defense Evasion is replaced by the two v19 tactics
    Stealth and Defense Impairment, taking STATUS/COVERAGE from 14 to 15
    tactic column pairs.

All user-input cells (sources[available], the detections table, and
techniques[detection rules modifier]) are left empty: this is a blank
template, not a filled-in assessment.
"""
import copy
import csv
import os
import sys

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
if os.path.exists(os.path.join(HERE, 'techniques.csv')):
    CSV_DIR = HERE
    ROOT = os.path.dirname(HERE)
elif os.path.exists(os.path.join(HERE, '20260809', 'techniques.csv')):
    CSV_DIR = os.path.join(HERE, '20260809')
    ROOT = HERE
else:
    CSV_DIR = HERE
    ROOT = HERE

BASE_2022 = os.path.join(ROOT, '20220505', 'AttackCoverage.xlsx')
BASE = BASE_2022 if os.path.exists(BASE_2022) else os.path.join(ROOT, 'AttackCoverage.xlsx')
OUT = os.path.join(ROOT, 'AttackCoverage.xlsx')

FILE_VERSION = 20260823
TEMPLATE_VERSION = 20260823
ATTACK_VERSION = 'v19.2 August 2026'

# Kill-chain order taken from the x-mitre-matrix object of enterprise-attack.json
# (v19.2).  Display labels keep the upstream spelling of "Command & Control".
ORDERED_TACTICS = [
    ('Reconnaissance', 'reconnaissance'),
    ('Resource Development', 'resource-development'),
    ('Initial Access', 'initial-access'),
    ('Execution', 'execution'),
    ('Persistence', 'persistence'),
    ('Privilege Escalation', 'privilege-escalation'),
    ('Stealth', 'stealth'),
    ('Defense Impairment', 'defense-impairment'),
    ('Credential Access', 'credential-access'),
    ('Discovery', 'discovery'),
    ('Lateral Movement', 'lateral-movement'),
    ('Collection', 'collection'),
    ('Command & Control', 'command-and-control'),
    ('Exfiltration', 'exfiltration'),
    ('Impact', 'impact'),
]

# STATUS / COVERAGE keep the upstream extent of 200 rows.
PAIR_MAX_ROW = 200

# --------------------------------------------------------------------------
# techniques sheet formulas, copied verbatim out of the upstream template.
# {ROW} is the only substitution; it appears solely in column I, which is the
# one formula upstream anchors with an explicit cell reference ($C2).
# --------------------------------------------------------------------------
TECH_FORMULAS = {
    'G': (True,
          'SUM(IF(ISERROR(FIND(sources[data source],techniques[[#This Row],[data sources]])),0,1)'
          '*IF(sources[available]="yes",1,0))'),
    'H': (True,
          'IF(techniques[[#This Row],[technique]]=techniques[[#This Row],[id]],'
          'SUM(IF(techniques[[#This Row],[id]]=techniques[technique],1,0))-1,"")'),
    'I': (True,
          'IF(SUM(IF($C{ROW}=detections[],IF(detections[is active]="yes",1,0),0))>0,'
          'SUM(IF($C{ROW}=detections[],IF(detections[is active]="yes",1,0),0)),"")'),
    'J': (True,
          'IF(IF(ISNUMBER(techniques[[#This Row],[number of sub techniques]]),'
          'techniques[[#This Row],[number of sub techniques]]>0,FALSE),'
          'SUM(IF(techniques[[#This Row],[technique]]=techniques[technique],'
          'techniques[detection rules for techique],0))'
          '-IF(ISNUMBER(techniques[[#This Row],[detection rules for techique]]),'
          'techniques[[#This Row],[detection rules for techique]],0),"")'),
    'K': (False,
          'IF(AND(ISNUMBER(techniques[[#This Row],[number of sub techniques]]),'
          'techniques[[#This Row],[number of sub techniques]]>1),'
          'techniques[number of sub techniques],1)'),
    'M': (True,
          'techniques[[#This Row],[minimum detection rules]]'
          '+IF(IF(ISNUMBER(techniques[[#This Row],[number of sub techniques]]),'
          'techniques[[#This Row],[number of sub techniques]]>0,FALSE),'
          'SUM(IF(techniques[[#This Row],[technique]]=techniques[technique],'
          'techniques[detection rules modifier],0)),'
          'techniques[[#This Row],[detection rules modifier]])'),
    'N': (False,
          'IF(ISNUMBER(techniques[[#This Row],[detection rules for techique]]),'
          'techniques[[#This Row],[detection rules for techique]],0)'
          '+IF(ISNUMBER(techniques[[#This Row],[detection rules for subtech]]),'
          'techniques[[#This Row],[detection rules for subtech]],0)'),
    'O': (False,
          'IF(techniques[[#This Row],[detection rules]]>='
          'techniques[[#This Row],[expected detection rules]],1,'
          'techniques[[#This Row],[detection rules]]/'
          'techniques[[#This Row],[expected detection rules]])'),
    'P': (False,
          'IF(techniques[[#This Row],[expected detection rules]]<=0,"disabled",'
          'IF(techniques[[#This Row],[data source available]]>0, '
          'IF(techniques[[#This Row],[detection rules]]>0,"detect","no detect"),'
          'IF(techniques[[#This Row],[detection rules]]>0,"inconsistent","no sources")))'),
    'R': (False,
          'IF(ISERROR(SEARCH(techniques[[#This Row],[id]],techniques[[#This Row],[name]])),"ERROR","")'),
    'S': (False,
          'IF(techniques[[#This Row],[expected detection rules]]<0,'
          '"expected detection rules negative!",'
          'IF(techniques[[#This Row],[detection rules]]>'
          'techniques[[#This Row],[expected detection rules]],'
          '"more detection rules than expected!",'
          'IF(AND(IF(ISNUMBER(techniques[[#This Row],[detection rules for techique]]),'
          'techniques[[#This Row],[detection rules for techique]]>0,FALSE),'
          'IF(ISNUMBER(techniques[[#This Row],[number of sub techniques]]),'
          'techniques[[#This Row],[number of sub techniques]]>0,FALSE)),'
          'IF(techniques[[#This Row],[detection rules for techique]] > '
          'IF(ISNUMBER(techniques[[#This Row],[detection rules modifier]]),'
          'techniques[[#This Row],[detection rules modifier]],0), '
          '"modifier should be increased for the technique", ""),"")))'),
}

STATUS_VALUE_FORMULA = (
    'IF({N}{R}<>"",IF(INDEX(techniques[detection rules for techique],'
    'MATCH({N}{R},techniques[name],0))>=1,'
    'INDEX(techniques[detection rules for techique],MATCH({N}{R},techniques[name],0)),""),"")')
COVERAGE_VALUE_FORMULA = (
    'IF({N}{R}<>"",IF(INDEX(techniques[coverage],MATCH({N}{R},techniques[name],0))>0,'
    'INDEX(techniques[coverage],MATCH({N}{R},techniques[name],0)),""),"")')
COVERAGE_TOTAL_FORMULA = 'SUM(IF({V}2:{V}%d<>"",{V}2:{V}%d,0))/SUM(IF({N}2:{N}%d<>"",1,0))'


def read_csvs():
    base = CSV_DIR
    with open(os.path.join(base, 'techniques.csv'), newline='', encoding='utf-8') as fh:
        techniques = list(csv.DictReader(fh))
    with open(os.path.join(base, 'tactics.csv'), newline='', encoding='utf-8') as fh:
        tactics = list(csv.DictReader(fh))
    with open(os.path.join(base, 'data_sources.csv'), newline='', encoding='utf-8') as fh:
        sources = [row[0].strip() for row in csv.reader(fh)][1:]
    sources = [s for s in sources if s]
    return techniques, tactics, sources


def grab_row_styles(ws, row, ncols):
    return {c: copy.copy(ws.cell(row, c)._style) for c in range(1, ncols + 1)}


def write_row(ws, row, styles, values):
    """values: dict of column index -> value (None clears)."""
    for col, style in styles.items():
        cell = ws.cell(row, col)
        cell.value = values.get(col)
        cell._style = copy.copy(style)


def blank_rows(ws, first, last, ncols):
    for row in range(first, last + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.style = 'Normal'


def build_techniques(ws, techniques):
    styles = grab_row_styles(ws, 2, 19)
    for offset, rec in enumerate(techniques):
        row = 2 + offset
        values = {
            1: rec['technique'],
            2: rec['id'],
            3: rec['name'],
            4: rec['tactics'],
            5: rec['data_sources'],
            6: int(rec['data_sources_num']),
        }
        for letter, (is_array, text) in TECH_FORMULAS.items():
            col = ord(letter) - 64
            formula = '=' + text.replace('{ROW}', str(row))
            values[col] = ArrayFormula('%s%d' % (letter, row), formula) if is_array else formula
        write_row(ws, row, styles, values)
    return 1 + len(techniques)


def build_tactics(ws, tactics):
    styles = grab_row_styles(ws, 2, 4)
    for offset, rec in enumerate(tactics):
        write_row(ws, 2 + offset, styles, {
            1: rec['name'], 2: rec['technique'],
            3: rec['technique_id'], 4: rec['technique_name'],
        })
    return 1 + len(tactics)


def build_sources(ws, sources):
    styles = grab_row_styles(ws, 2, 3)
    for offset, name in enumerate(sources):
        write_row(ws, 2 + offset, styles, {1: name})
    return 1 + len(sources)


def build_pair_sheet(ws, kind, per_tactic):
    """Rewrite a STATUS/COVERAGE sheet with one column pair per tactic."""
    name_styles = {r: copy.copy(ws.cell(r, 1)._style) for r in range(1, PAIR_MAX_ROW + 1)}
    value_styles = {r: copy.copy(ws.cell(r, 2)._style) for r in range(1, PAIR_MAX_ROW + 1)}

    for index, (label, slug) in enumerate(ORDERED_TACTICS):
        ncol = index * 2 + 1
        vcol = ncol + 1
        nl = get_column_letter(ncol)
        vl = get_column_letter(vcol)

        hdr = ws.cell(1, ncol)
        hdr.value = label
        hdr._style = copy.copy(name_styles[1])
        total = ws.cell(1, vcol)
        if kind == 'STATUS':
            total.value = '=SUM(%s2:%s%d)' % (vl, vl, PAIR_MAX_ROW)
        else:
            text = COVERAGE_TOTAL_FORMULA % (PAIR_MAX_ROW, PAIR_MAX_ROW, PAIR_MAX_ROW)
            total.value = ArrayFormula('%s1' % vl, '=' + text.format(V=vl, N=nl))
        total._style = copy.copy(value_styles[1])

        names = per_tactic[slug]
        for row in range(2, PAIR_MAX_ROW + 1):
            ncell = ws.cell(row, ncol)
            ncell.value = names[row - 2] if row - 2 < len(names) else None
            ncell._style = copy.copy(name_styles[row])
            template = STATUS_VALUE_FORMULA if kind == 'STATUS' else COVERAGE_VALUE_FORMULA
            vcell = ws.cell(row, vcol)
            vcell.value = ArrayFormula('%s%d' % (vl, row),
                                       '=' + template.format(N=nl, R=row))
            vcell._style = copy.copy(value_styles[row])

        ws.column_dimensions[nl].width = 20.0
        ws.column_dimensions[vl].width = 6.5 if kind == 'COVERAGE' else 5.0

    for src, dst in (('AA', 'AC'), ('AB', 'AD')):
        ref = ws.column_dimensions[src]
        new = ws.column_dimensions[dst]
        new.width = ref.width
        new.bestFit = ref.bestFit


def rebuild_pair_cf(ws, ntactics):
    """Re-emit the sheet's CF blocks across ntactics column pairs.

    Each upstream block covers either every name column or every value column.
    Regenerating from the block's parity also normalises the upstream anomaly
    where STATUS's "no detect"/"disabled" blocks started at C6/D6 instead of
    row 2, which left Resource Development rows 2-5 uncoloured.
    """
    saved = []
    for cf in ws.conditional_formatting:
        ranges = sorted(cf.sqref.ranges, key=lambda r: (r.min_col, r.min_row))
        parity = ranges[0].min_col % 2
        start = min(r.min_row for r in ranges)
        saved.append((parity, start, list(cf.rules)))

    ws.conditional_formatting = ConditionalFormattingList()
    for parity, start, rules in saved:
        cols = [i * 2 + (1 if parity else 2) for i in range(ntactics)]
        sqref = ' '.join('%s%d:%s%d' % (get_column_letter(c), start,
                                        get_column_letter(c), PAIR_MAX_ROW)
                         for c in cols)
        for rule in rules:
            ws.conditional_formatting.add(sqref, rule)


def rebuild_block_cf(ws, ref):
    saved = [list(cf.rules) for cf in ws.conditional_formatting]
    ws.conditional_formatting = ConditionalFormattingList()
    for rules in saved:
        for rule in rules:
            ws.conditional_formatting.add(ref, rule)


def retarget_table(ws, name, ref):
    table = ws._tables[name]
    table.ref = ref
    if table.autoFilter is not None:
        table.autoFilter.ref = ref
    if table.sortState is not None:
        table.sortState = None
    ncols = len(table.tableColumns)
    width = _ref_width(ref)
    if ncols != width:
        raise SystemExit('table %s: %d tableColumns but ref %s is %d wide'
                         % (name, ncols, ref, width))


def _ref_width(ref):
    from openpyxl.utils import range_boundaries
    min_col, _, max_col, _ = range_boundaries(ref)
    return max_col - min_col + 1


def main():
    techniques, tactics, sources = read_csvs()

    all_by_tactic = {slug: [] for _, slug in ORDERED_TACTICS}
    parents_by_tactic = {slug: [] for _, slug in ORDERED_TACTICS}
    for rec in techniques:
        for slug in (s.strip() for s in rec['tactics'].split('|')):
            if slug not in all_by_tactic:
                raise SystemExit('unknown tactic slug %r in techniques.csv' % slug)
            all_by_tactic[slug].append(rec['name'])
            if rec['technique'] == rec['id']:
                parents_by_tactic[slug].append(rec['name'])
    worst = max(len(v) for v in all_by_tactic.values())
    if worst > PAIR_MAX_ROW - 1:
        raise SystemExit('a tactic has %d techniques, more than the %d STATUS rows'
                         % (worst, PAIR_MAX_ROW - 1))

    wb = load_workbook(BASE)

    tech_last = build_techniques(wb['techniques'], techniques)
    tact_last = build_tactics(wb['tactics'], tactics)
    src_last = build_sources(wb['sources'], sources)
    blank_rows(wb['sources'], src_last + 1, 110, 3)
    blank_rows(wb['detections'], 2, 999, 8)

    build_pair_sheet(wb['STATUS'], 'STATUS', all_by_tactic)
    build_pair_sheet(wb['COVERAGE'], 'COVERAGE', parents_by_tactic)

    retarget_table(wb['techniques'], 'techniques', 'A1:P%d' % tech_last)
    retarget_table(wb['tactics'], 'tactics', 'A1:D%d' % tact_last)
    retarget_table(wb['sources'], 'sources', 'A1:C%d' % src_last)
    retarget_table(wb['detections'], 'detections', 'A1:H999')

    wb['techniques'].auto_filter.ref = 'Q1:S%d' % tech_last
    rebuild_block_cf(wb['techniques'], 'A2:P%d' % tech_last)
    rebuild_block_cf(wb['sources'], 'A2:C%d' % src_last)
    rebuild_pair_cf(wb['STATUS'], len(ORDERED_TACTICS))
    rebuild_pair_cf(wb['COVERAGE'], len(ORDERED_TACTICS))

    ver = wb['version']
    ver['A1'], ver['B1'] = 'Current Excel file version', FILE_VERSION
    ver['A2'], ver['B2'] = 'Based on template version', TEMPLATE_VERSION
    ver['A3'], ver['B3'] = 'Base on MITRE ATT&CK® version', ATTACK_VERSION

    wb.calculation.fullCalcOnLoad = True
    wb.save(OUT)

    print('wrote %s' % OUT)
    print('  techniques rows 2..%d (%d)' % (tech_last, len(techniques)))
    print('  tactics    rows 2..%d (%d)' % (tact_last, len(tactics)))
    print('  sources    rows 2..%d (%d)' % (src_last, len(sources)))
    print('  detections empty, table A1:H999')
    print('  %d tactic pairs, A..%s' % (len(ORDERED_TACTICS),
                                        get_column_letter(len(ORDERED_TACTICS) * 2)))
    for label, slug in ORDERED_TACTICS:
        print('    %-22s STATUS=%3d COVERAGE=%3d'
              % (label, len(all_by_tactic[slug]), len(parents_by_tactic[slug])))


if __name__ == '__main__':
    sys.exit(main())
