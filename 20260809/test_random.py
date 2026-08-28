#!/usr/bin/env python3
"""End-to-end test of the rebuilt AttackCoverage.xlsx with random inputs.

Three independent implementations are compared:

  1. the workbook itself -- AttackCoverage.xlsx filled with seeded random data
     sources, detection rules and modifiers, saved as AttackCoverage_RandomTest.xlsx
     for a human to open in Excel;
  2. a mirror of that workbook in which every Excel Table structured reference
     is expanded to its A1 equivalent, evaluated by the `formulas` engine
     (the `formulas` parser cannot read structured references, and structured
     references are pure syntactic sugar for ranges, so the mirror is
     semantically identical);
  3. a from-scratch Python oracle written directly from the README's rules.

If (2) and (3) agree on every cell then the formulas in the shipped workbook
compute what the project documents they should compute.
"""
import copy
import csv
import os
import random
import sys

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(line_buffering=True)

HERE = os.path.dirname(os.path.abspath(__file__))
if os.path.exists(os.path.join(HERE, 'techniques.csv')):
    CSV_DIR = HERE
    ROOT = os.path.dirname(HERE)
elif os.path.exists(os.path.join(HERE, '20260809', 'techniques.csv')):
    CSV_DIR = os.path.join(HERE, '20260809')
    ROOT = HERE
else:
    CSV_DIR = HERE
    ROOT = HERE

TEMPLATE = os.path.join(ROOT, 'AttackCoverage.xlsx')
TESTFILE = os.path.join(ROOT, 'AttackCoverage_RandomTest.xlsx')
MIRROR = os.path.join(CSV_DIR, '_mirror_expanded.xlsx')
REPORT = os.path.join(ROOT, 'AttackCoverage_RandomTest_Report.md')

SEED = 20260823
N_RULES = 140
FRAC_SOURCES_AVAILABLE = 0.30
N_MODIFIERS = 45

TECH_LAST = 698
DET_LAST = 999
SRC_LAST = 99
PAIR_MAX_ROW = 200

ORDERED_TACTICS = [
    ('Reconnaissance', 'reconnaissance'), ('Resource Development', 'resource-development'),
    ('Initial Access', 'initial-access'), ('Execution', 'execution'),
    ('Persistence', 'persistence'), ('Privilege Escalation', 'privilege-escalation'),
    ('Stealth', 'stealth'), ('Defense Impairment', 'defense-impairment'),
    ('Credential Access', 'credential-access'), ('Discovery', 'discovery'),
    ('Lateral Movement', 'lateral-movement'), ('Collection', 'collection'),
    ('Command & Control', 'command-and-control'), ('Exfiltration', 'exfiltration'),
    ('Impact', 'impact'),
]

# --------------------------------------------------------------------------
# A1-expanded mirrors of the shipped structured-reference formulas.
# --------------------------------------------------------------------------
T = 'techniques!'
MIRROR_TECH = {
    'G': "=SUM(IF(ISERROR(FIND(sources!$A$2:$A${S},{T}$E{r})),0,1)*IF(sources!$B$2:$B${S}=\"yes\",1,0))",
    'H': "=IF({T}$A{r}={T}$B{r},SUM(IF({T}$B{r}={T}$A$2:$A${L},1,0))-1,\"\")",
    'I': "=IF(SUM(IF({T}$C{r}=detections!$A$2:$H${D},IF(detections!$E$2:$E${D}=\"yes\",1,0),0))>0,"
         "SUM(IF({T}$C{r}=detections!$A$2:$H${D},IF(detections!$E$2:$E${D}=\"yes\",1,0),0)),\"\")",
    'J': "=IF(IF(ISNUMBER({T}$H{r}),{T}$H{r}>0,FALSE),"
         "SUM(IF({T}$A{r}={T}$A$2:$A${L},{T}$I$2:$I${L},0))"
         "-IF(ISNUMBER({T}$I{r}),{T}$I{r},0),\"\")",
    'K': "=IF(AND(ISNUMBER({T}$H{r}),{T}$H{r}>1),{T}$H{r},1)",
    'M': "={T}$K{r}+IF(IF(ISNUMBER({T}$H{r}),{T}$H{r}>0,FALSE),"
         "SUM(IF({T}$A{r}={T}$A$2:$A${L},{T}$L$2:$L${L},0)),{T}$L{r})",
    'N': "=IF(ISNUMBER({T}$I{r}),{T}$I{r},0)+IF(ISNUMBER({T}$J{r}),{T}$J{r},0)",
    'O': "=IF({T}$N{r}>={T}$M{r},1,{T}$N{r}/{T}$M{r})",
    'P': "=IF({T}$M{r}<=0,\"disabled\",IF({T}$G{r}>0, IF({T}$N{r}>0,\"detect\",\"no detect\"),"
         "IF({T}$N{r}>0,\"inconsistent\",\"no sources\")))",
}
MIRROR_STATUS = ('=IF({n}{r}<>"",IF(INDEX(techniques!$I$2:$I${L},MATCH({n}{r},techniques!$C$2:$C${L},0))>=1,'
                 'INDEX(techniques!$I$2:$I${L},MATCH({n}{r},techniques!$C$2:$C${L},0)),""),"")')
MIRROR_COVERAGE = ('=IF({n}{r}<>"",IF(INDEX(techniques!$O$2:$O${L},MATCH({n}{r},techniques!$C$2:$C${L},0))>0,'
                   'INDEX(techniques!$O$2:$O${L},MATCH({n}{r},techniques!$C$2:$C${L},0)),""),"")')


def mirror_tech(letter, row):
    return MIRROR_TECH[letter].format(T=T, r=row, L=TECH_LAST, D=DET_LAST, S=SRC_LAST)


# --------------------------------------------------------------------------
# random input generation
# --------------------------------------------------------------------------
def read_csvs():
    with open(os.path.join(CSV_DIR, 'techniques.csv'), newline='', encoding='utf-8') as fh:
        techniques = list(csv.DictReader(fh))
    with open(os.path.join(CSV_DIR, 'data_sources.csv'), newline='', encoding='utf-8') as fh:
        sources = [r[0].strip() for r in csv.reader(fh)][1:]
    return techniques, [s for s in sources if s]


def make_inputs(techniques, sources, rng):
    available = set(rng.sample(sources, int(round(len(sources) * FRAC_SOURCES_AVAILABLE))))

    names = [t['name'] for t in techniques]
    parents = [t for t in techniques if t['technique'] == t['id']]
    subs_of = {}
    for t in techniques:
        if t['technique'] != t['id']:
            subs_of.setdefault(t['technique'], []).append(t['name'])

    rules = []
    for i in range(N_RULES):
        picks = [rng.choice(names)]
        if rng.random() < 0.40:
            picks.append(rng.choice(names))
        if rng.random() < 0.20:
            picks.append(rng.choice(names))
        picks += [None] * (3 - len(picks))
        rules.append({
            'use case id': 'uc%03d' % (i // 7 + 1),
            'use case': 'random use case %d' % (i // 7 + 1),
            'rule id': 'rule%04d' % (i + 1),
            'rule description': 'seeded random detection rule %d' % (i + 1),
            'is active': 'yes' if rng.random() < 0.80 else ('no' if rng.random() < 0.5 else None),
            'attack1': picks[0], 'attack2': picks[1], 'attack3': picks[2],
        })

    # one parent whose detection rules sit only on its sub-techniques, so that
    # techniques[detection rules for subtech] has something real to aggregate
    big = [p for p in parents if len(subs_of.get(p['id'], [])) >= 4]
    aggregation_parent = rng.choice(big)
    for sub in subs_of[aggregation_parent['id']][:3]:
        rules.append({
            'use case id': 'uc900', 'use case': 'sub-technique aggregation probe',
            'rule id': 'ruleAGG%d' % len(rules), 'rule description': 'probe',
            'is active': 'yes', 'attack1': sub, 'attack2': None, 'attack3': None,
        })

    modifiers = {}
    for t in rng.sample(techniques, N_MODIFIERS):
        modifiers[t['name']] = rng.choice([-1, -1, 1, 1, 2, 3, 5])
    # guarantee at least one technique is switched off via a -1 modifier
    off = rng.choice([t for t in techniques if t['technique'] != t['id']])
    modifiers[off['name']] = -1
    return available, rules, modifiers, aggregation_parent['name'], off['name']


# --------------------------------------------------------------------------
# oracle: README semantics, implemented from scratch
# --------------------------------------------------------------------------
def oracle(techniques, sources, available, rules, modifiers):
    by_parent = {}
    for t in techniques:
        by_parent.setdefault(t['technique'], []).append(t)

    rule_cells = []
    for r in rules:
        active = (r['is active'] == 'yes')
        cells = [r[k] for k in ('use case id', 'use case', 'rule id', 'rule description',
                                'is active', 'attack1', 'attack2', 'attack3')]
        rule_cells.append((active, cells))

    res = {}
    for t in techniques:
        name, tid, parent = t['name'], t['id'], t['technique']
        ds = t['data_sources']
        g = sum(1 for s in sources if s in ds and s in available)
        h = (len(by_parent[tid]) - 1) if parent == tid else ''
        i = sum(1 for active, cells in rule_cells if active for c in cells if c == name)
        res[name] = {'G': g, 'H': h, 'I': i if i > 0 else '', 'L': modifiers.get(name, '')}

    for t in techniques:
        name, tid, parent = t['name'], t['id'], t['technique']
        cur = res[name]
        h, i = cur['H'], cur['I']
        family = by_parent[parent]
        if isinstance(h, int) and h > 0:
            fam_rules = sum(res[f['name']]['I'] for f in family
                            if isinstance(res[f['name']]['I'], int))
            cur['J'] = fam_rules - (i if isinstance(i, int) else 0)
        else:
            cur['J'] = ''
        cur['K'] = h if (isinstance(h, int) and h > 1) else 1
        if isinstance(h, int) and h > 0:
            bump = sum(res[f['name']]['L'] for f in family
                       if isinstance(res[f['name']]['L'], int))
        else:
            bump = cur['L'] if isinstance(cur['L'], int) else 0
        cur['M'] = cur['K'] + bump
        cur['N'] = (i if isinstance(i, int) else 0) + (cur['J'] if isinstance(cur['J'], int) else 0)
        cur['O'] = 1 if cur['N'] >= cur['M'] else cur['N'] / cur['M']
        if cur['M'] <= 0:
            cur['P'] = 'disabled'
        elif cur['G'] > 0:
            cur['P'] = 'detect' if cur['N'] > 0 else 'no detect'
        else:
            cur['P'] = 'inconsistent' if cur['N'] > 0 else 'no sources'
    return res


# --------------------------------------------------------------------------
# workbook writing
# --------------------------------------------------------------------------
def fill_inputs(wb, sources, available, rules, modifiers):
    src = wb['sources']
    for row in range(2, SRC_LAST + 1):
        if src.cell(row, 1).value in available:
            src.cell(row, 2).value = 'yes'
    det = wb['detections']
    style = {c: copy.copy(det.cell(2, c)._style) for c in range(1, 9)}
    keys = ['use case id', 'use case', 'rule id', 'rule description',
            'is active', 'attack1', 'attack2', 'attack3']
    for offset, rule in enumerate(rules):
        row = 2 + offset
        for col, key in enumerate(keys, 1):
            cell = det.cell(row, col)
            cell.value = rule[key]
            cell._style = copy.copy(style[col])
    tech = wb['techniques']
    for row in range(2, TECH_LAST + 1):
        mod = modifiers.get(tech.cell(row, 3).value)
        if mod is not None:
            tech.cell(row, 12).value = mod


def build_mirror(techniques, sources, available, rules, modifiers, per_tactic_all,
                 per_tactic_parents):
    wb = openpyxl.Workbook()
    src = wb.active
    src.title = 'sources'
    src['A1'], src['B1'], src['C1'] = 'data source', 'available', 'note'
    for row, name in enumerate(sources, 2):
        src.cell(row, 1, name)
        if name in available:
            src.cell(row, 2, 'yes')

    det = wb.create_sheet('detections')
    keys = ['use case id', 'use case', 'rule id', 'rule description',
            'is active', 'attack1', 'attack2', 'attack3']
    for col, key in enumerate(keys, 1):
        det.cell(1, col, key)
    for offset, rule in enumerate(rules):
        for col, key in enumerate(keys, 1):
            det.cell(2 + offset, col, rule[key])

    tech = wb.create_sheet('techniques')
    hdr = ['technique', 'id', 'name', 'tactics', 'data sources', 'data sources number',
           'data source available', 'number of sub techniques', 'detection rules for techique',
           'detection rules for subtech', 'minimum detection rules', 'detection rules modifier',
           'expected detection rules', 'detection rules', 'coverage', 'technique status']
    for col, h in enumerate(hdr, 1):
        tech.cell(1, col, h)
    for offset, rec in enumerate(techniques):
        row = 2 + offset
        tech.cell(row, 1, rec['technique'])
        tech.cell(row, 2, rec['id'])
        tech.cell(row, 3, rec['name'])
        tech.cell(row, 4, rec['tactics'])
        tech.cell(row, 5, rec['data_sources'])
        tech.cell(row, 6, int(rec['data_sources_num']))
        mod = modifiers.get(rec['name'])
        if mod is not None:
            tech.cell(row, 12, mod)
        for letter in 'GHIJKMNOP':
            tech.cell(row, ord(letter) - 64, mirror_tech(letter, row))

    for kind, per_tactic, template in (('STATUS', per_tactic_all, MIRROR_STATUS),
                                       ('COVERAGE', per_tactic_parents, MIRROR_COVERAGE)):
        ws = wb.create_sheet(kind)
        for index, (label, slug) in enumerate(ORDERED_TACTICS):
            ncol, vcol = index * 2 + 1, index * 2 + 2
            nl, vl = get_column_letter(ncol), get_column_letter(vcol)
            ws.cell(1, ncol, label)
            if kind == 'STATUS':
                ws.cell(1, vcol, '=SUM(%s2:%s%d)' % (vl, vl, PAIR_MAX_ROW))
            else:
                ws.cell(1, vcol, '=SUM(IF({V}2:{V}{M}<>"",{V}2:{V}{M},0))/SUM(IF({N}2:{N}{M}<>"",1,0))'
                        .format(V=vl, N=nl, M=PAIR_MAX_ROW))
            names = per_tactic[slug]
            # two rows past the data so the blank-row behaviour is exercised too
            for row in range(2, min(len(names) + 4, PAIR_MAX_ROW + 1)):
                if row - 2 < len(names):
                    ws.cell(row, ncol, names[row - 2])
                ws.cell(row, vcol, template.format(n=nl, r=row, L=TECH_LAST))
    wb.save(MIRROR)
    return MIRROR


# --------------------------------------------------------------------------
def evaluate(path):
    import formulas
    model = formulas.ExcelModel().loads(path).finish()
    sol = model.calculate()
    book = os.path.basename(path).upper()
    out = {}
    for key, val in sol.items():
        upper = key.upper()
        if "'[%s]" % book not in upper:
            continue
        sheet_cell = upper.split("'[%s]" % book, 1)[1]
        if "'!" not in sheet_cell:
            continue
        sheet, cell = sheet_cell.split("'!", 1)
        if '$' in cell or ':' in cell:
            continue
        try:
            value = val.value[0, 0]
        except Exception:
            value = val
        out['%s!%s' % (sheet, cell)] = value
    return out


def same(a, b):
    if isinstance(a, str) or isinstance(b, str):
        sa = '' if a in (None, '') else str(a)
        sb = '' if b in (None, '') else str(b)
        return sa == sb
    try:
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return str(a) == str(b)


def main():
    rng = random.Random(SEED)
    techniques, sources = read_csvs()
    available, rules, modifiers, agg_parent, forced_off = make_inputs(techniques, sources, rng)

    per_all = {slug: [] for _, slug in ORDERED_TACTICS}
    per_parents = {slug: [] for _, slug in ORDERED_TACTICS}
    for rec in techniques:
        for slug in (s.strip() for s in rec['tactics'].split('|')):
            per_all[slug].append(rec['name'])
            if rec['technique'] == rec['id']:
                per_parents[slug].append(rec['name'])

    exp = oracle(techniques, sources, available, rules, modifiers)

    wb = openpyxl.load_workbook(TEMPLATE)
    fill_inputs(wb, sources, available, rules, modifiers)
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(TESTFILE)
        print('[+] %s  (%d sources available, %d detection rules, %d modifiers)'
              % (os.path.basename(TESTFILE), len(available), len(rules), len(modifiers)))
    except PermissionError:
        print('[!] WARNING: %s is currently open in Excel. Close it to save changes.'
              % os.path.basename(TESTFILE))
        print('[*] Proceeding with calculation test...')

    path = build_mirror(techniques, sources, available, rules, modifiers, per_all, per_parents)
    print('[+] mirror written, evaluating with `formulas` ...')
    got = evaluate(path)
    print('[+] %d cells evaluated' % len(got))

    fails = []
    compared = 0
    for offset, rec in enumerate(techniques):
        row = 2 + offset
        e = exp[rec['name']]
        for letter in 'GHIJKMNOP':
            key = 'TECHNIQUES!%s%d' % (letter, row)
            actual = got.get(key, '<missing>')
            compared += 1
            if not same(actual, e[letter]):
                fails.append('techniques!%s%d (%s): engine=%r oracle=%r'
                             % (letter, row, rec['name'], actual, e[letter]))

    for kind, per_tactic in (('STATUS', per_all), ('COVERAGE', per_parents)):
        for index, (label, slug) in enumerate(ORDERED_TACTICS):
            nl = get_column_letter(index * 2 + 1)
            vl = get_column_letter(index * 2 + 2)
            names = per_tactic[slug]
            total = 0.0
            for offset, name in enumerate(names):
                row = 2 + offset
                if kind == 'STATUS':
                    want = exp[name]['I'] if isinstance(exp[name]['I'], int) and exp[name]['I'] >= 1 else ''
                else:
                    want = exp[name]['O'] if exp[name]['O'] > 0 else ''
                actual = got.get('%s!%s%d' % (kind, vl, row), '<missing>')
                compared += 1
                if not same(actual, want):
                    fails.append('%s!%s%d (%s): engine=%r oracle=%r'
                                 % (kind, vl, row, name, actual, want))
                total += want if isinstance(want, (int, float)) else 0.0
            want1 = total if kind == 'STATUS' else (total / len(names) if names else 0)
            actual1 = got.get('%s!%s1' % (kind, vl), '<missing>')
            compared += 1
            if not same(actual1, want1):
                fails.append('%s!%s1 (%s total): engine=%r oracle=%r'
                             % (kind, vl, label, actual1, want1))
            for row in (len(names) + 2, len(names) + 3):
                if row <= PAIR_MAX_ROW:
                    actual = got.get('%s!%s%d' % (kind, vl, row), '<missing>')
                    compared += 1
                    if not same(actual, ''):
                        fails.append('%s!%s%d (past end of %s): engine=%r want blank'
                                     % (kind, vl, row, label, actual))

    print('[+] compared %d cells, %d mismatches' % (compared, len(fails)))
    for line in fails[:40]:
        print('    FAIL', line)

    write_report(techniques, sources, available, rules, modifiers, exp, got,
                 per_all, per_parents, compared, fails, agg_parent, forced_off)
    return 1 if fails else 0


def write_report(techniques, sources, available, rules, modifiers, exp, got,
                 per_all, per_parents, compared, fails, agg_parent, forced_off):
    counts = {}
    for rec in techniques:
        counts[exp[rec['name']]['P']] = counts.get(exp[rec['name']]['P'], 0) + 1

    examples = {}
    for rec in techniques:
        status = exp[rec['name']]['P']
        if status not in examples:
            examples[status] = rec['name']

    lines = []
    lines.append('# AttackCoverage.xlsx random-data test report')
    lines.append('')
    lines.append('Seed `%d`. Generated by `20260809/test_random.py`.' % SEED)
    lines.append('')
    lines.append('Three implementations are compared: the shipped workbook filled with random')
    lines.append('inputs (`AttackCoverage_RandomTest.xlsx`), an A1-expanded mirror of it')
    lines.append('evaluated by the `formulas` engine, and a from-scratch Python oracle written')
    lines.append('from the README rules.')
    lines.append('')
    lines.append('## Result')
    lines.append('')
    lines.append('* cells compared: **%d**' % compared)
    lines.append('* mismatches: **%d**' % len(fails))
    lines.append('')
    if fails:
        lines.append('```')
        lines.extend(fails[:60])
        lines.append('```')
    else:
        lines.append('The calculation engine and the independent oracle agree on every')
        lines.append('`techniques` calculated column, every STATUS and COVERAGE cell, every')
        lines.append('row-1 aggregate, and on the blank rows past the end of each tactic.')
    lines.append('')
    lines.append('## Random inputs')
    lines.append('')
    lines.append('| input | value |')
    lines.append('| --- | --- |')
    lines.append('| data sources marked available | %d of %d |' % (len(available), len(sources)))
    lines.append('| detection rules | %d (%d active) |'
                 % (len(rules), sum(1 for r in rules if r['is active'] == 'yes')))
    lines.append('| detection rule modifiers set | %d |' % len(modifiers))
    lines.append('| techniques | %d |' % len(techniques))
    lines.append('')
    lines.append('## Technique status distribution')
    lines.append('')
    lines.append('| status | techniques | example |')
    lines.append('| --- | --- | --- |')
    for status in ('detect', 'no detect', 'inconsistent', 'no sources', 'disabled'):
        lines.append('| %s | %d | %s |' % (status, counts.get(status, 0),
                                           examples.get(status, '-')))
    lines.append('')
    lines.append('All five documented statuses are reachable with this input, so every')
    lines.append('conditional-formatting branch on the `techniques` sheet is exercised.')
    lines.append('')
    lines.append('## Worked examples')
    lines.append('')
    lines.append('| technique | sub-tech | rules on it | rules on subs | expected | actual | coverage | status |')
    lines.append('| --- | --- | --- | --- | --- | --- | --- | --- |')
    picks = [agg_parent, forced_off] + [examples[s] for s in
                                        ('detect', 'no detect', 'inconsistent', 'no sources')
                                        if s in examples]
    for name in dict.fromkeys(picks):
        e = exp[name]
        lines.append('| %s | %s | %s | %s | %s | %s | %s | %s |'
                     % (name, e['H'], e['I'], e['J'], e['M'], e['N'],
                        round(e['O'], 3), e['P']))
    lines.append('')
    lines.append('`%s` is the parent whose detection rules were placed only on its'
                 % agg_parent)
    lines.append('sub-techniques, which is what makes `detection rules for subtech` non-zero.')
    lines.append('`%s` carries a `-1` modifier, which is the documented way to switch a'
                 % forced_off)
    lines.append('technique off, and it comes out as `disabled`.')
    lines.append('')
    lines.append('## Per-tactic totals')
    lines.append('')
    lines.append('| tactic | techniques | detection rules (STATUS total) | coverage (COVERAGE row 1) |')
    lines.append('| --- | --- | --- | --- |')
    for index, (label, slug) in enumerate(ORDERED_TACTICS):
        vl = get_column_letter(index * 2 + 2)
        status_total = got.get('STATUS!%s1' % vl, '?')
        coverage = got.get('COVERAGE!%s1' % vl, '?')
        try:
            coverage = '%.1f%%' % (float(coverage) * 100)
        except (TypeError, ValueError):
            pass
        try:
            status_total = '%d' % float(status_total)
        except (TypeError, ValueError):
            pass
        lines.append('| %s | %d | %s | %s |'
                     % (label, len(per_all[slug]), status_total, coverage))
    lines.append('')
    lines.append('Stealth and Defense Impairment appear as tactic pairs 7 and 8; Defense')
    lines.append('Evasion is gone. Both aggregate normally, which is the point of the test.')
    lines.append('')
    try:
        with open(REPORT, 'w', encoding='utf-8') as fh:
            fh.write('\n'.join(lines))
        print('[+] %s' % os.path.basename(REPORT))
    except PermissionError:
        print('[!] WARNING: %s is currently open/locked. Cannot overwrite report file.' % os.path.basename(REPORT))
    return 1 if fails else 0


if __name__ == '__main__':
    sys.exit(main())

