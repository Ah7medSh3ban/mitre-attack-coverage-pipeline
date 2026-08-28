#!/usr/bin/env python3
"""Static integrity checks for a built AttackCoverage.xlsx.

Verifies the things that actually break silently: table metadata consistency,
conditional-formatting coverage, structured-reference targets, and the
techniques[name] join key that every other sheet depends on.
"""
import os
import re
import sys
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

NTACTICS = 15
PAIR_MAX_ROW = 200

problems = []
notes = []


def check(cond, msg):
    if cond:
        notes.append('  OK   %s' % msg)
    else:
        problems.append('  FAIL %s' % msg)


def formula_text(value):
    if isinstance(value, ArrayFormula):
        return value.text
    return value if isinstance(value, str) and value.startswith('=') else None


def main(path):
    wb = load_workbook(path)
    check(wb.sheetnames == ['COVERAGE', 'STATUS', 'detections', 'techniques',
                            'tactics', 'sources', 'version'],
          'sheet order preserved: %s' % wb.sheetnames)

    # ---- tables -----------------------------------------------------------
    tables = {}
    for ws in wb.worksheets:
        for table in ws._tables.values():
            tables[table.name] = (ws.title, table)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            check(len(table.tableColumns) == max_col - min_col + 1,
                  'table %s: %d columns == ref %s width'
                  % (table.name, len(table.tableColumns), table.ref))
            check(min_row == 1, 'table %s starts at row 1' % table.name)
    check(set(tables) == {'techniques', 'detections', 'tactics', 'sources'},
          'all four tables present: %s' % sorted(tables))

    tech = wb['techniques']
    tech_tbl = tables['techniques'][1]
    tech_last = range_boundaries(tech_tbl.ref)[3]
    tech_cols = [c.name for c in tech_tbl.tableColumns]
    check(tech_cols[5] == 'data sources\nnumber',
          'header newline preserved in %r' % tech_cols[5])
    check('detection rules for techique' in tech_cols,
          'upstream header typo preserved')

    # ---- structured reference targets exist -------------------------------
    valid = set()
    for name, (_, table) in tables.items():
        valid.add(name.lower())
        for col in table.tableColumns:
            valid.add('%s[%s]' % (name.lower(), col.name.lower()))
    pattern = re.compile(r'(\w+)\[\[#This Row\],\[([^\]]*)\]\]|(\w+)\[([^\[\]]*)\]')
    unresolved = set()
    scanned = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                text = formula_text(cell.value)
                if not text:
                    continue
                scanned += 1
                for m in pattern.finditer(text):
                    tbl = (m.group(1) or m.group(3)).lower()
                    col = (m.group(2) if m.group(1) else m.group(4)) or ''
                    if col in ('', '#This Row'):
                        key = tbl
                    else:
                        key = '%s[%s]' % (tbl, col.lower())
                    if key not in valid:
                        unresolved.add('%s!%s -> %s' % (ws.title, cell.coordinate, key))
    check(not unresolved, 'all structured refs resolve (%d formulas scanned)%s'
          % (scanned, '' if not unresolved else '; bad: %s' % sorted(unresolved)[:5]))

    # ---- array formula self-reference -------------------------------------
    bad_ref = [(ws.title, c.coordinate) for ws in wb.worksheets for r in ws.iter_rows()
               for c in r if isinstance(c.value, ArrayFormula) and c.value.ref != c.coordinate]
    check(not bad_ref, 'every ArrayFormula ref equals its own cell%s'
          % ('' if not bad_ref else ' (bad: %s)' % bad_ref[:5]))

    # ---- techniques join key ---------------------------------------------
    names, ids = [], []
    for row in range(2, tech_last + 1):
        names.append(tech.cell(row, 3).value)
        ids.append(tech.cell(row, 2).value)
    check(all(names) and len(set(names)) == len(names),
          '%d technique names, all unique (join key safe)' % len(names))
    check(len(set(ids)) == len(ids), '%d technique ids, all unique' % len(ids))
    bad_pair = [n for n, i in zip(names, ids) if not n.endswith('(%s)' % i)]
    check(not bad_pair, 'every name ends with its own id%s'
          % ('' if not bad_pair else ' (bad: %s)' % bad_pair[:3]))
    nameset = set(names)

    # ---- conditional formatting ------------------------------------------
    def cf_columns(ws):
        cols = {}
        for cf in ws.conditional_formatting:
            for rng in cf.sqref.ranges:
                for col in range(rng.min_col, rng.max_col + 1):
                    lo, hi = cols.get(col, (10 ** 9, 0))
                    cols[col] = (min(lo, rng.min_row), max(hi, rng.max_row))
        return cols

    cols = cf_columns(tech)
    check(sorted(cols) == list(range(1, 17)),
          'techniques CF spans columns A:P (%s)' % sorted(cols))
    check(all(v == (2, tech_last) for v in cols.values()),
          'techniques CF spans rows 2..%d (%s)' % (tech_last, set(cols.values())))

    src_last = range_boundaries(tables['sources'][1].ref)[3]
    cols = cf_columns(wb['sources'])
    check(sorted(cols) == [1, 2, 3] and all(v == (2, src_last) for v in cols.values()),
          'sources CF is A2:C%d (%s)' % (src_last, set(cols.values())))

    for sheet in ('STATUS', 'COVERAGE'):
        ws = wb[sheet]
        cols = cf_columns(ws)
        want = list(range(1, NTACTICS * 2 + 1))
        check(sorted(cols) == want,
              '%s CF spans all %d tactic pairs A:%s (missing %s)'
              % (sheet, NTACTICS, get_column_letter(NTACTICS * 2),
                 sorted(set(want) - set(cols)) or 'none'))
        starts = {c: v[0] for c, v in cols.items()}
        expect = 1 if sheet == 'COVERAGE' else 2
        check(set(starts.values()) == {expect} or (sheet == 'COVERAGE' and set(starts.values()) == {1, 2}),
              '%s CF start rows %s' % (sheet, sorted(set(starts.values()))))
        check(all(v[1] == PAIR_MAX_ROW for v in cols.values()),
              '%s CF ends at row %d' % (sheet, PAIR_MAX_ROW))

    # ---- STATUS / COVERAGE content ---------------------------------------
    tech_by_name = {tech.cell(r, 3).value: r for r in range(2, tech_last + 1)}
    parents = {tech.cell(r, 3).value for r in range(2, tech_last + 1)
               if tech.cell(r, 1).value == tech.cell(r, 2).value}
    for sheet in ('STATUS', 'COVERAGE'):
        ws = wb[sheet]
        orphans, misplaced, holes = [], [], []
        for i in range(NTACTICS):
            ncol, vcol = i * 2 + 1, i * 2 + 2
            nl = get_column_letter(ncol)
            vl = get_column_letter(vcol)
            seen_blank = False
            for row in range(2, PAIR_MAX_ROW + 1):
                val = ws.cell(row, ncol).value
                if val in (None, ''):
                    seen_blank = True
                else:
                    if seen_blank:
                        holes.append('%s!%s%d' % (sheet, nl, row))
                    if val not in nameset:
                        orphans.append('%s!%s%d=%r' % (sheet, nl, row, val))
                    elif sheet == 'COVERAGE' and val not in parents:
                        misplaced.append('%s!%s%d=%r' % (sheet, nl, row, val))
                text = formula_text(ws.cell(row, vcol).value) or ''
                if '%s%d' % (nl, row) not in text:
                    misplaced.append('%s!%s%d formula does not read %s%d'
                                     % (sheet, vl, row, nl, row))
        check(not orphans, '%s: every technique name exists in techniques[name]%s'
              % (sheet, '' if not orphans else ' (orphans: %s)' % orphans[:3]))
        check(not holes, '%s: no gaps in the technique name columns%s'
              % (sheet, '' if not holes else ' (%s)' % holes[:3]))
        check(not misplaced, '%s: value formulas point at their own name column%s'
              % (sheet, '' if not misplaced else ' (%s)' % misplaced[:3]))
        labels = [ws.cell(1, i * 2 + 1).value for i in range(NTACTICS)]
        check('Stealth' in labels and 'Defense Impairment' in labels
              and 'Defense Evasion' not in labels,
              '%s headers: Stealth + Defense Impairment present, Defense Evasion gone' % sheet)
        check(labels.index('Stealth') == 6 and labels.index('Defense Impairment') == 7,
              '%s: Stealth is pair 7, Defense Impairment is pair 8' % sheet)

    # ---- template must ship empty ----------------------------------------
    det = wb['detections']
    filled = [(r, c) for r in range(2, 1000) for c in range(1, 9)
              if det.cell(r, c).value not in (None, '')]
    check(not filled, 'detections table is empty (%d filled cells)' % len(filled))
    src = wb['sources']
    check(not [r for r in range(2, src_last + 1) if src.cell(r, 2).value not in (None, '')],
          'no data source pre-marked available')
    check(not [r for r in range(2, tech_last + 1) if tech.cell(r, 12).value not in (None, '')],
          'no detection rules modifier pre-set')

    # ---- leftovers from the smaller/larger v11 extents --------------------
    check(not [r for r in range(src_last + 1, 111) if src.cell(r, 1).value not in (None, '')],
          'no leftover data source rows past %d' % src_last)

    # ---- raw xml sanity ---------------------------------------------------
    with zipfile.ZipFile(path) as zf:
        wanted = ['xl/theme/theme1.xml', 'xl/styles.xml']
        for entry in wanted:
            check(entry in zf.namelist(), '%s present' % entry)
        styles = zf.read('xl/styles.xml').decode('utf-8')
        check(styles.count('<dxf>') == 72, 'all 72 differential styles kept (%d)'
              % styles.count('<dxf>'))
        arrays = sum(zf.read(n).decode('utf-8').count('<calculatedColumnFormula array="1"')
                     for n in zf.namelist() if n.startswith('xl/tables/'))
        check(arrays == 5, '5 array calculated-column formulas kept (%d)' % arrays)
        for name in zf.namelist():
            if name.startswith('xl/worksheets/sheet'):
                body = zf.read(name).decode('utf-8')
                check('#This Row' not in body or '[[#This Row],[' in body,
                      '%s: this-row refs intact' % name)

    ver = wb['version']
    check(str(ver['B3'].value).startswith('v19'),
          'version sheet says %r' % ver['B3'].value)

    print('\n'.join(notes))
    if problems:
        print('\n%d PROBLEM(S):' % len(problems))
        print('\n'.join(problems))
        return 1
    print('\nALL %d CHECKS PASSED' % len(notes))
    return 0


if __name__ == '__main__':
    if len(sys.argv) > 1:
        target = sys.argv[1]
    else:
        here = os.path.dirname(os.path.abspath(__file__))
        root = here if os.path.exists(os.path.join(here, 'AttackCoverage.xlsx')) else os.path.dirname(here)
        target = os.path.join(root, 'AttackCoverage.xlsx')
    sys.exit(main(target))
